import tkinter as tk
import tkinter.ttk as ttk
from tkinter import messagebox, filedialog
import os, shutil
import random
import string
from encryptor import Encryptor
from decryptor import Decryptor
import pyAesCrypt
from openpyxl import *
import tempfile
from tkinter import *
from Gdrive import *
import subprocess


class HomePage:

	def __init__( self, root, master, service, home, user, userId, password, dirlists, filelists ):
		# build ui
		self.root = root
		self.mainframe = master
		self.service = service
		self.home = os.path.join( home, user )
		self.encdir = os.path.join( 'ENCDIR', user )
		print( self.encdir )
		self.user = user
		self.userId = userId
		self.password = password
		# self.dirname = dirname
		self.dirlists = dirlists
		self.filelists = filelists
		# master = root
		self.show_F = ttk.Frame( master )
		self.fileList_LB = tk.Listbox( self.show_F )
		self.fileList_LB.config( font='{Arial} 12 {bold}', height='12', relief='flat', selectmode='single' )
		self.fileList_LB.config( state='normal', takefocus=True, width='10' )
		# self.fileList_LB.place( anchor='nw', x='0', y='0' )
		self.fileList_LB.pack( expand='true', fill='both', side='left' )
		self.scrollbar_V = ttk.Scrollbar( self.show_F )
		self.scrollbar_V.config( orient='vertical' )
		self.scrollbar_V.pack( expand='true', fill='y', side='top' )
		self.scrollbar_V.config( command=self.fileList_LB.yview )
		self.fileList_LB.configure( yscrollcommand=self.scrollbar_V.set )
		self.fileList_LB.bind( '<Double-Button>', self.openFolder )
		self.fileList_LB.bind( '<Return>', self.openFolder )
		self.fileList_LB.bind( '<KP_Enter>', self.openFolder )
		self.fileList_LB.bind( '<F5>', self.insertFolders )
		self.fileList_LB.bind( '<F2>', self.renameFolder )

		self.button_1 = ttk.Button( self.show_F )
		self.button_1.config( text='Open' )
		self.button_1.place( anchor='n', relx='0.10', rely='0.92', x='0', y='0' )
		self.button_1.configure( command=self.openFolder, width='8' )

		self.button_2 = ttk.Button( self.show_F )
		self.button_2.config( text='Create' )
		self.button_2.place( anchor='n', relx='0.29', rely='0.92', x='0', y='0' )
		self.button_2.configure( command=self.createFolder, width='8' )

		self.button_3 = ttk.Button( self.show_F )
		self.button_3.config( text='Delete' )
		self.button_3.place( anchor='n', relx='0.67', rely='0.92', x='0', y='0' )
		self.button_3.configure( command=self.deleteFolder, width='8' )

		self.button_4 = ttk.Button( self.show_F )
		self.button_4.config( text='Logout' )
		self.button_4.place( anchor='n', relx='0.86', rely='0.92', x='0', y='0' )
		self.button_4.configure( command=self.logout, width='8' )

		self.button_5 = ttk.Button( self.show_F )
		self.button_5.place( anchor='n', relx='0.48', rely='0.92', x='0', y='0' )
		self.button_5.configure( text='Rename', command=self.renameFolder, width='7' )

		self.show_F.config( height='500', relief='flat', width='480' )
		self.show_F.place( anchor='nw', height='500', width='480', x='0', y='0' )

		# Main widget
		self.mainwindow = self.show_F
		# for line in range( 100 ):
		# 	self.fileList_LB.insert( END, "This is line number " + str( line ) )
		self.config()
		self.login()
		self.insertFolders()
		self.fileList_LB.focus_set()
		#print( self.home )

	def random_str( self, n ):
		return ( ''.join( random.choices( string.ascii_uppercase + string.digits + string.ascii_lowercase, k=n ) ) )

	def config( self ):
		self.root.protocol( "WM_DELETE_WINDOW", self.EXIT )
		if not os.path.exists( self.home ):
			os.makedirs( self.home )
		if not os.path.exists( self.encdir ):
			os.makedirs( self.encdir )
		# self.filepath = ""
		# self.home = os.path.abspath( os.path.join( os.path.expanduser( '~' ), "AppData\\Local\\Programs\\Invisible" ) )
		# self.userdir = os.path.join( self.home, 'Users' )
		self.bufferSize = 64 * 1024
		self.s_w = self.root.winfo_screenwidth()
		self.s_h = self.root.winfo_screenheight()
		self.wow = 480
		self.how = 530
		x_c = ( ( self.s_w / 2 ) - ( self.wow / 2 ) )
		y_c = ( ( self.s_h / 2 ) - ( self.how / 2 ) )
		self.root.geometry( "%dx%d+%d+%d" % ( self.wow, self.how, x_c, y_c ) )
		self.mainframe.config( text=self.user )
		self.mainframe.update()
		self.dirlistsID = getFileId( self.service, self.dirlists, self.userId )
		if self.dirlistsID == None:
			tempxl = os.path.join( tempfile.gettempdir(), self.random_str( 12 ) + '.xlsx' )
			WB = Workbook()
			sheet = WB.active

			sheet.cell( row=1, column=1 ).value = "Folder Id"
			sheet.cell( row=1, column=2 ).value = "Folder Name"
			sheet.cell( row=1, column=3 ).value = "Temp Name"

			WB.save( tempxl )
			pyAesCrypt.encryptFile( tempxl, os.path.join( self.home, self.dirlists ), self.password, self.bufferSize )
			WB.close()
			os.remove( tempxl )
			self.dirlistsID = uploadFile(
			    self.service, self.dirlists, os.path.join( self.home, self.dirlists ), self.userId
			    )

	def login( self ):
		DE = Decryptor( self.encdir, self.dirlists, self.filelists, self.password )
		DE.decrypt()

	def reconfig( self ):

		self.button_1.configure( command=self.openFile, width='8' )
		self.button_2.configure( text='Add File', command=self.addFile, width='8' )
		self.button_3.configure( command=self.deleteFile, width='8' )
		self.button_4.configure( text='Back', command=self.back, width='8' )
		self.button_5.configure( text='Update', command=self.updateFile, width='8' )
		self.fileList_LB.bind( '<Double-Button>', self.openFile )
		self.fileList_LB.bind( '<Return>', self.openFile )
		self.fileList_LB.bind( '<KP_Enter>', self.openFile )
		self.fileList_LB.bind( '<F5>', self.insertFiles )
		self.fileList_LB.bind( '<F2>', self.renameFile )
		self.fileList_LB.focus_set()

		pass

	def insertFolders( self, event=None ):
		try:
			self.fileList_LB.delete( 0, END )
			self.mainframe.update()

			downloadFile( self.service, self.dirlistsID, os.path.join( self.home, self.dirlists ) )
			tempxl = os.path.join( tempfile.gettempdir(), self.random_str( 12 ) + '.xlsx' )
			self.dirlist = []
			if os.path.exists( os.path.join( self.home, self.dirlists ) ):
				pyAesCrypt.decryptFile(
				    os.path.join( self.home, self.dirlists ), tempxl, self.password, self.bufferSize
				    )
				WB = load_workbook( tempxl )
				sheet = WB.active
				L_R = sheet.max_row
				# print( L_R )
				for i in range( 2, L_R + 1 ):
					self.fileList_LB.insert( END, sheet.cell( row=i, column=2 ).value )
					self.dirlist.append(
					    [
					        sheet.cell( row=i, column=1 ).value,
					        sheet.cell( row=i, column=2 ).value,
					        sheet.cell( row=i, column=3 ).value
					        ]
					    )
					#print( sheet.cell( row=i, column=1 ).value )
				# root_dir = sheet.cell( row=dest_dir, column=2 ).value
				WB.close()
				os.remove( tempxl )
				os.remove( os.path.join( self.home, self.dirlists ) )
			self.fileList_LB.focus_set()
			self.fileList_LB.selection_set( 0 )
			# print( self.dirlist )
		except Exception as e:
			messagebox.showerror( "Error", e )

	def insertFiles( self, event=None ):
		try:
			self.fileList_LB.delete( 0, END )
			self.mainframe.update()

			downloadFile( self.service, self.filelistsID, os.path.join( self.home, self.filelists ) )
			tempxl = os.path.join( tempfile.gettempdir(), self.random_str( 12 ) + '.xlsx' )
			self.filelist = []
			if os.path.exists( os.path.join( self.home, self.filelists ) ):
				pyAesCrypt.decryptFile(
				    os.path.join( self.home, self.filelists ), tempxl, self.password, self.bufferSize
				    )
				WB = load_workbook( tempxl )
				sheet = WB.active
				L_R = sheet.max_row
				# print( L_R )
				for i in range( 2, L_R + 1 ):
					# detail = "{:_<30}{:_>20}".format(
					#     sheet.cell( row=i, column=2 ).value)

					# print( sheet.cell( row=i, column=1 ).value )
					self.fileList_LB.insert( END, sheet.cell( row=i, column=2 ).value )
					self.filelist.append(
					    [
					        sheet.cell( row=i, column=1 ).value,
					        sheet.cell( row=i, column=2 ).value,
					        sheet.cell( row=i, column=3 ).value
					        ]
					    )
					#print( sheet.cell( row=i, column=1 ).value )
				# root_dir = sheet.cell( row=dest_dir, column=2 ).value
				WB.close()
				os.remove( tempxl )
				os.remove( os.path.join( self.home, self.filelists ) )
			self.fileList_LB.focus_set()
			self.fileList_LB.selection_set( 0 )
			# print( self.filelist )
		except Exception as e:
			messagebox.showerror( "Error", e )

	def openFile( self, event=None ):
		try:
			selected = self.fileList_LB.curselection()
			if selected != ():
				self.mainframe.config( text="Opening..." )
				self.mainframe.update()
				fileId = self.filelist[ selected[ 0 ] ][ 0 ]
				filename = self.filelist[ selected[ 0 ] ][ 1 ]
				filetempname = self.filelist[ selected[ 0 ] ][ 2 ]
				tempxl = os.path.join( tempfile.gettempdir(), self.random_str( 12 ) )
				if os.path.exists( os.path.join( self.home, filename ) ):
					flag = 'no'
					flag = messagebox.askquestion(
					    "Exist",
					    "Yes : To open Drive file, it will replace Local PC file. \n No : To open Local PC file."
					    )
					if flag == 'yes':
						os.system( "attrib -s -h \"{}\"".format( os.path.join( self.home, filename ) ) )
						downloadFile( self.service, fileId, tempxl )

						pyAesCrypt.decryptFile(
						    tempxl, os.path.join( self.home, filename ), self.password, self.bufferSize
						    )
						os.system( "attrib +s +h \"{}\"".format( os.path.join( self.home, filename ) ) )

						openfile = subprocess.Popen(
						    [ "start", " ", os.path.abspath( os.path.join( self.home, filename ) ) ], shell=True
						    )
						os.remove( tempxl )
					else:
						openfile = subprocess.Popen(
						    [ "start", " ", os.path.abspath( os.path.join( self.home, filename ) ) ], shell=True
						    )

				else:
					downloadFile( self.service, fileId, tempxl )
					pyAesCrypt.decryptFile(
					    tempxl, os.path.join( self.home, filename ), self.password, self.bufferSize
					    )
					os.system( "attrib +s +h \"{}\"".format( os.path.join( self.home, filename ) ) )

					openfile = subprocess.Popen(
					    [ "start", " ", os.path.abspath( os.path.join( self.home, filename ) ) ], shell=True
					    )
					os.remove( tempxl )

			else:
				messagebox.showerror( 'Invalid', "Invalid Selection !" )
		except Exception as e:
			print( e )
			messagebox.showerror( "Error", e )
		finally:
			self.mainframe.config( text=self.user )
			self.mainframe.update()

	def addFile( self ):
		try:
			File = filedialog.askopenfilename( title="Select File" )
			Name = os.path.basename( File ).replace( " ", "_" )
			# print( self.folderId )

			# print( Name )
			if File != "":
				self.mainframe.config( text="Uploading..." )
				self.mainframe.update()
				random_name = self.random_str( 15 )
				pyAesCrypt.encryptFile( File, os.path.join( self.home, random_name ), self.password, self.bufferSize )
				fileId = uploadFile( self.service, random_name, os.path.join( self.home, random_name ), self.folderId )
				downloadFile( self.service, self.filelistsID, os.path.join( self.home, self.filelists ) )
				tempxl = os.path.join( tempfile.gettempdir(), self.random_str( 12 ) + '.xlsx' )
				# print( tempxl )
				pyAesCrypt.decryptFile(
				    os.path.join( self.home, self.filelists ), tempxl, self.password, self.bufferSize
				    )
				WB = load_workbook( tempxl )
				sheet = WB.active
				L_R = sheet.max_row + 1
				# print( L_R )
				sheet.cell( row=L_R, column=1 ).value = fileId
				sheet.cell( row=L_R, column=2 ).value = Name
				sheet.cell( row=L_R, column=3 ).value = random_name
				WB.save( tempxl )

				pyAesCrypt.encryptFile(
				    tempxl, os.path.join( self.home, self.filelists ), self.password, self.bufferSize
				    )
				updateFile( self.service, self.filelistsID, self.filelists, os.path.join( self.home, self.filelists ) )

				os.remove( tempxl )
				os.remove( os.path.join( self.home, random_name ) )
				try:
					os.remove( File )
				except:
					pass
				self.insertFiles()
				self.mainframe.config( text=self.user )
				self.mainframe.update()
				messagebox.showinfo( 'Done', 'File Encrypted & Uploaded !' )
		except Exception as e:
			messagebox.showerror( "Error", e )
		finally:
			self.mainframe.config( text=self.user )
			self.mainframe.update()

	def updateFile( self ):
		try:
			selected = self.fileList_LB.curselection()
			if selected != ():
				self.mainframe.config( text="Updating..." )
				self.mainframe.update()
				fileId = self.filelist[ selected[ 0 ] ][ 0 ]
				filename = self.filelist[ selected[ 0 ] ][ 1 ]
				filetempname = self.filelist[ selected[ 0 ] ][ 2 ]
				tempxl = os.path.join( tempfile.gettempdir(), self.random_str( 12 ) + '.xlsx' )
				if os.path.exists( os.path.join( self.home, filename ) ):
					flag = 'no'
					flag = messagebox.askquestion( "Update", "Are you sure, you want to update ?" )
					if flag == 'yes':

						pyAesCrypt.encryptFile(
						    os.path.join( self.home, filename ), tempxl, self.password, self.bufferSize
						    )
						updateFile( self.service, fileId, filetempname, tempxl )
					os.remove( tempxl )
					try:
						os.remove( os.path.join( self.home, filename ) )
					except:
						pass
					self.insertFiles()
					messagebox.showinfo( 'Done', 'File Updated !' )

				else:
					messagebox.showerror( 'Error', "No Update Found !" )

				self.mainframe.config( text=self.user )
				self.mainframe.update()
			else:
				messagebox.showerror( 'Invalid', "Invalid Selection !" )
		except Exception as e:
			messagebox.showerror( "Error", e )
		finally:
			self.mainframe.config( text=self.user )
			self.mainframe.update()

	def renameFile( self, event=None ):

		def Add( event=None ):
			try:
				Name = name_E.get()
				if Name != '':
					self.mainframe.config( text="Updating..." )
					self.mainframe.update()
					downloadFile( self.service, self.filelistsID, os.path.join( self.home, self.filelists ) )
					tempxl = os.path.join( tempfile.gettempdir(), self.random_str( 12 ) + '.xlsx' )
					# print( tempxl )
					pyAesCrypt.decryptFile(
					    os.path.join( self.home, self.filelists ), tempxl, self.password, self.bufferSize
					    )
					WB = load_workbook( tempxl )
					sheet = WB.active
					L_R = selected[ 0 ] + 2
					# print( L_R )
					# sheet.cell( row=L_R, column=1 ).value = folderID
					sheet.cell( row=L_R, column=2 ).value = Name.replace( " ", "_" )
					# sheet.cell( row=L_R, column=3 ).value = random_name
					WB.save( tempxl )

					pyAesCrypt.encryptFile(
					    tempxl, os.path.join( self.home, self.filelists ), self.password, self.bufferSize
					    )
					updateFile(
					    self.service, self.filelistsID, self.filelists, os.path.join( self.home, self.filelists )
					    )
					os.remove( tempxl )
					if os.path.exists( os.path.join( self.home, filename ) ):
						os.rename( os.path.join( self.home, filename ), os.path.join( self.home, Name ) )
					frame_1.destroy()
					self.insertFiles()
					self.mainframe.config( text=self.user )
					self.mainframe.update()
				else:
					messagebox.showerror( 'Empty', 'Empty field!' )
			except Exception as e:
				messagebox.showerror( "Error", e )
			finally:
				self.mainframe.config( text=self.user )
				self.mainframe.update()

		def cancel():
			frame_1.destroy()

		try:
			selected = self.fileList_LB.curselection()
			if selected != ():
				folderId = self.filelist[ selected[ 0 ] ][ 0 ]
				filename = self.filelist[ selected[ 0 ] ][ 1 ]
				frame_1 = ttk.Frame( self.show_F )
				folderName_L = ttk.Label( frame_1 )
				folderName_L.config( font='{Arial} 14 {bold}', takefocus=False, text='Name' )
				folderName_L.grid( padx='5' )
				name_E = ttk.Entry( frame_1 )
				name_E.config( font='{Arial} 12 {bold}', takefocus=True, width='20' )
				name_E.grid( column='1', padx='5', pady='5', row='0' )
				name_E.bind( '<Return>', Add )
				name_E.bind( '<KP_Enter>', Add )

				name_E.delete( 0, END )
				name_E.insert( 0, filename )
				add_B = ttk.Button( frame_1 )
				add_B.config( text='Rename' )
				add_B.grid( column='0', pady='10', row='2' )
				add_B.configure( command=Add )
				cancel_B = ttk.Button( frame_1 )
				cancel_B.config( takefocus=True, text='Cancel' )
				cancel_B.grid( column='2', padx='10', pady='10', row='2' )
				cancel_B.configure( command=cancel )

				frame_1.config( height='200', width='480', relief='raised' )
				frame_1.place( anchor='n', relx='0.48', rely='0.4', x='0', y='0' )
				frame_1.pack_propagate( 0 )
				name_E.focus_set()
		except Exception as e:
			messagebox.showerror( "Error", e )

	def deleteFile( self ):
		try:
			selected = self.fileList_LB.curselection()
			if selected != ():
				fileId = self.filelist[ selected[ 0 ] ][ 0 ]
				filename = self.filelist[ selected[ 0 ] ][ 1 ]
				flag = 'no'
				flag = messagebox.askquestion( "Hide", "Are you sure you want to permanently delete this file ?" )
				if flag == 'yes':
					self.mainframe.config( text="Deleting..." )
					self.mainframe.update()
					downloadFile( self.service, self.filelistsID, os.path.join( self.home, self.filelists ) )
					tempxl = os.path.join( tempfile.gettempdir(), self.random_str( 12 ) + '.xlsx' )
					pyAesCrypt.decryptFile(
					    os.path.join( self.home, self.filelists ), tempxl, self.password, self.bufferSize
					    )
					WB = load_workbook( tempxl )
					# print( tempxl )
					sheet = WB.active
					sheet.delete_rows( selected[ 0 ] + 2 )
					WB.save( tempxl )
					WB.close()
					pyAesCrypt.encryptFile(
					    tempxl, os.path.join( self.home, self.filelists ), self.password, self.bufferSize
					    )
					os.remove( tempxl )
					updateFile(
					    self.service, self.filelistsID, self.filelists, os.path.join( self.home, self.filelists )
					    )
					deleteFile( self.service, fileId )
					if os.path.exists( os.path.join( self.home, filename ) ):
						os.remove( os.path.join( self.home, filename ) )
					self.insertFiles()
					messagebox.showinfo( 'Done', 'File Deleted !' )
					self.mainframe.config( text=self.user )
					self.mainframe.update()
			else:
				messagebox.showerror( 'Invalid', "Invalid Selection !" )
		except Exception as e:
			messagebox.showerror( "Error", e )
		finally:
			self.mainframe.config( text=self.user )
			self.mainframe.update()

	def back( self ):

		self.button_1.configure( command=self.openFolder, width='8' )
		self.button_2.configure( text='Create', command=self.createFolder, width='8' )
		self.button_3.configure( command=self.deleteFolder, width='8' )
		self.button_4.configure( text='Logout', command=self.logout, width='8' )
		self.button_5.configure( text='Rename', command=self.renameFolder, width='8' )
		self.fileList_LB.bind( '<Double-Button>', self.openFolder )
		self.fileList_LB.bind( '<Return>', self.openFolder )
		self.fileList_LB.bind( '<KP_Enter>', self.openFolder )
		self.fileList_LB.bind( '<F5>', self.insertFolders )
		self.fileList_LB.bind( '<F2>', self.renameFolder )

		self.show_F.update()
		self.insertFolders()

	def createFolder( self ):

		def Add( event=None ):
			Name = name_E.get()
			if Name != '':
				self.mainframe.config( text="Creating..." )
				self.mainframe.update()
				random_name = self.random_str( 12 )
				folderID = createFolder( self.service, random_name, self.userId )
				downloadFile( self.service, self.dirlistsID, os.path.join( self.home, self.dirlists ) )
				tempxl = os.path.join( tempfile.gettempdir(), self.random_str( 12 ) + '.xlsx' )
				# print( tempxl )
				pyAesCrypt.decryptFile(
				    os.path.join( self.home, self.dirlists ), tempxl, self.password, self.bufferSize
				    )
				WB = load_workbook( tempxl )
				sheet = WB.active
				L_R = sheet.max_row + 1
				# print( L_R )
				sheet.cell( row=L_R, column=1 ).value = folderID
				sheet.cell( row=L_R, column=2 ).value = Name
				sheet.cell( row=L_R, column=3 ).value = random_name
				WB.save( tempxl )

				pyAesCrypt.encryptFile(
				    tempxl, os.path.join( self.home, self.dirlists ), self.password, self.bufferSize
				    )
				updateFile( self.service, self.dirlistsID, self.dirlists, os.path.join( self.home, self.dirlists ) )
				os.remove( tempxl )
				frame_1.destroy()
				self.insertFolders()
				self.mainframe.config( text=self.user )
				self.mainframe.update()
			else:
				messagebox.showerror( 'Empty', 'Empty field!' )

		def cancel():
			frame_1.destroy()

		try:
			frame_1 = ttk.Frame( self.show_F )
			folderName_L = ttk.Label( frame_1 )
			folderName_L.config( font='{Arial} 14 {bold}', takefocus=False, text='Name' )
			folderName_L.grid( padx='5' )
			name_E = ttk.Entry( frame_1 )
			name_E.config( font='{Arial} 12 {bold}', takefocus=True, width='20' )
			name_E.grid( column='1', padx='5', pady='5', row='0' )
			name_E.bind( '<Return>', Add )
			name_E.bind( '<KP_Enter>', Add )
			add_B = ttk.Button( frame_1 )
			add_B.config( text='Create' )
			add_B.grid( column='0', pady='10', row='2' )
			add_B.configure( command=Add )
			cancel_B = ttk.Button( frame_1 )
			cancel_B.config( takefocus=True, text='Cancel' )
			cancel_B.grid( column='2', padx='10', pady='10', row='2' )
			cancel_B.configure( command=cancel )

			frame_1.config( height='200', width='480', relief='raised' )
			frame_1.place( anchor='n', relx='0.48', rely='0.4', x='0', y='0' )
			frame_1.pack_propagate( 0 )
			name_E.focus_set()
		except Exception as e:
			messagebox.showerror( "Error", e )

	def openFolder( self, event=None ):
		try:
			selected = self.fileList_LB.curselection()
			if selected != ():
				self.mainframe.config( text="Opening..." )
				self.mainframe.update()
				self.reconfig()
				self.folderId = self.dirlist[ selected[ 0 ] ][ 0 ]
				# print( self.folderId )
				self.filelistsID = getFileId( self.service, self.filelists, self.folderId )

				if self.filelistsID == None:
					tempxl = os.path.join( tempfile.gettempdir(), self.random_str( 12 ) + '.xlsx' )
					WB = Workbook()
					sheet = WB.active

					sheet.cell( row=1, column=1 ).value = "File ID"
					sheet.cell( row=1, column=2 ).value = "File Name"
					sheet.cell( row=1, column=3 ).value = "Temp Name"

					WB.save( tempxl )
					pyAesCrypt.encryptFile(
					    tempxl, os.path.join( self.home, self.filelists ), self.password, self.bufferSize
					    )
					WB.close()
					os.remove( tempxl )
					self.filelistsID = uploadFile(
					    self.service, self.filelists, os.path.join( self.home, self.filelists ), self.folderId
					    )
				self.insertFiles()
				self.mainframe.config( text=self.user )
				self.mainframe.update()
			else:
				messagebox.showerror( 'Invalid', "Invalid Selection !" )
		except Exception as e:
			messagebox.showerror( "Error", e )
		finally:
			self.mainframe.config( text=self.user )
			self.mainframe.update()

	def deleteFolder( self ):
		try:
			selected = self.fileList_LB.curselection()
			if selected != ():
				folderId = self.dirlist[ selected[ 0 ] ][ 0 ]

				flag = 'no'
				flag = messagebox.askquestion( "Hide", "Are you sure you want to permanently delete this folder ?" )
				if flag == 'yes':
					self.mainframe.config( text="Deleting..." )
					self.mainframe.update()
					downloadFile( self.service, self.dirlistsID, os.path.join( self.home, self.dirlists ) )
					tempxl = os.path.join( tempfile.gettempdir(), self.random_str( 12 ) + '.xlsx' )
					pyAesCrypt.decryptFile(
					    os.path.join( self.home, self.dirlists ), tempxl, self.password, self.bufferSize
					    )
					WB = load_workbook( tempxl )
					# print( tempxl )
					sheet = WB.active
					sheet.delete_rows( selected[ 0 ] + 2 )
					WB.save( tempxl )
					WB.close()
					pyAesCrypt.encryptFile(
					    tempxl, os.path.join( self.home, self.dirlists ), self.password, self.bufferSize
					    )

					os.remove( tempxl )
					updateFile( self.service, self.dirlistsID, self.dirlists, os.path.join( self.home, self.dirlists ) )
					deleteFile( self.service, folderId )
					self.insertFolders()
					self.mainframe.config( text=self.user )
					self.mainframe.update()

			else:
				messagebox.showerror( 'Invalid', "Invalid Selection !" )
		except Exception as e:
			messagebox.showerror( "Error", e )
		finally:
			self.mainframe.config( text=self.user )
			self.mainframe.update()

	def renameFolder( self, event=None ):

		def Add( event=None ):
			Name = name_E.get()
			if Name != '':
				self.mainframe.config( text="Updating..." )
				self.mainframe.update()
				downloadFile( self.service, self.dirlistsID, os.path.join( self.home, self.dirlists ) )
				tempxl = os.path.join( tempfile.gettempdir(), self.random_str( 12 ) + '.xlsx' )
				# print( tempxl )
				pyAesCrypt.decryptFile(
				    os.path.join( self.home, self.dirlists ), tempxl, self.password, self.bufferSize
				    )
				WB = load_workbook( tempxl )
				sheet = WB.active
				L_R = selected[ 0 ] + 2
				# print( L_R )
				# sheet.cell( row=L_R, column=1 ).value = folderID
				sheet.cell( row=L_R, column=2 ).value = Name
				# sheet.cell( row=L_R, column=3 ).value = random_name
				WB.save( tempxl )

				pyAesCrypt.encryptFile(
				    tempxl, os.path.join( self.home, self.dirlists ), self.password, self.bufferSize
				    )
				updateFile( self.service, self.dirlistsID, self.dirlists, os.path.join( self.home, self.dirlists ) )
				os.remove( tempxl )
				frame_1.destroy()
				self.insertFolders()
				self.mainframe.config( text=self.user )
				self.mainframe.update()
			else:
				messagebox.showerror( 'Empty', 'Empty field!' )

		def cancel():
			frame_1.destroy()

		try:
			selected = self.fileList_LB.curselection()
			if selected != ():
				folderId = self.dirlist[ selected[ 0 ] ][ 0 ]
				foldername = self.dirlist[ selected[ 0 ] ][ 1 ]
				frame_1 = ttk.Frame( self.show_F )
				folderName_L = ttk.Label( frame_1 )
				folderName_L.config( font='{Arial} 14 {bold}', takefocus=False, text='Name' )
				folderName_L.grid( padx='5' )
				name_E = ttk.Entry( frame_1 )
				name_E.config( font='{Arial} 12 {bold}', takefocus=True, width='20' )
				name_E.grid( column='1', padx='5', pady='5', row='0' )
				name_E.bind( '<Return>', Add )
				name_E.bind( '<KP_Enter>', Add )
				name_E.delete( 0, END )
				name_E.insert( 0, foldername )
				add_B = ttk.Button( frame_1 )
				add_B.config( text='Rename' )
				add_B.grid( column='0', pady='10', row='2' )
				add_B.configure( command=Add )
				cancel_B = ttk.Button( frame_1 )
				cancel_B.config( takefocus=True, text='Cancel' )
				cancel_B.grid( column='2', padx='10', pady='10', row='2' )
				cancel_B.configure( command=cancel )

				frame_1.config( height='200', width='480', relief='raised' )
				frame_1.place( anchor='n', relx='0.48', rely='0.4', x='0', y='0' )
				frame_1.pack_propagate( 0 )
				name_E.focus_set()
		except Exception as e:
			messagebox.showerror( "Error", e )

	def logout( self ):
		EN = Encryptor( self.home, self.encdir, self.dirlists, self.filelists, self.password )
		EN.encrypt()
		self.root.protocol( "WM_DELETE_WINDOW", self.root.destroy )

		self.show_F.destroy()
		self.wow = 480
		self.how = 250
		x_c = ( ( self.s_w / 2 ) - ( self.wow / 2 ) )
		y_c = ( ( self.s_h / 2 ) - ( self.how / 2 ) )
		self.root.geometry( "%dx%d+%d+%d" % ( self.wow, self.how, x_c, y_c ) )
		self.mainframe.config( text="Personal" )
		self.mainframe.update()

	def EXIT( self ):
		EN = Encryptor( self.home, self.encdir, self.dirlists, self.filelists, self.password )
		EN.encrypt()
		self.root.destroy()

	def run( self ):
		self.mainwindow.mainloop()


if __name__ == '__main__':
	root = tk.Tk()
	app = HomePage( root )
	app.run()
