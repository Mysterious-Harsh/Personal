import os
import string
import random
import shutil
import pyAesCrypt
from openpyxl import *
import tempfile
from tkinter import messagebox


class Encryptor:

	def __init__( self, src_dir, dest_dir, dirlists, filelists, key ):
		self.key = key
		self.src_dir = src_dir
		self.dest_dir = dest_dir
		self.dirlists = dirlists
		self.filelists = filelists
		self.bufferSize = 64 * 1024

	def random_str( self, n ):
		return ( ''.join( random.choices( string.ascii_uppercase + string.digits + string.ascii_lowercase, k=n ) ) )

	def encrypt( self ):
		if not os.path.exists( self.dest_dir ):
			os.makedirs( self.dest_dir )

		root_dir = self.dest_dir

		FILEPATHXL = os.path.join( tempfile.gettempdir(), self.random_str( 12 ) + '.xlsx' )

		WB = Workbook()
		sheet = WB.active
		cell = 1
		for path, subdirs, files in os.walk( self.src_dir ):

			for name in files:
				temp = os.path.join( root_dir, self.random_str( 15 ) )
				while os.path.exists( temp ):
					temp = os.path.join( root_dir, self.random_str( 15 ) )
				filepath = os.path.relpath( os.path.join( path, name ) )
				# print( filepath )
				# print( temp )
				sheet.cell( row=cell, column=1 ).value = filepath
				sheet.cell( row=cell, column=2 ).value = temp
				if os.path.exists( temp ):
					os.system( "attrib -s -h  \"{}\"".format( temp ) )
				pyAesCrypt.encryptFile( filepath, temp, self.key, self.bufferSize )

				cell += 1
		for root, dirs, files in os.walk( self.src_dir ):
			for name in files:
				try:
					os.remove( os.path.join( root, name ) )
				except:
					pass
			for name in dirs:
				try:
					shutil.rmtree( os.path.join( root, name ) )
				except:
					pass
		try:
			shutil.rmtree( self.src_dir )
		except:
			pass
		WB.save( FILEPATHXL )
		if os.path.exists( os.path.join( root_dir, self.filelists ) ):
			os.system( "attrib -s -h  \"{}\"".format( os.path.join( root_dir, self.filelists ) ) )
		pyAesCrypt.encryptFile( FILEPATHXL, os.path.join( root_dir, self.filelists ), self.key, self.bufferSize )
		WB.close()
		os.remove( FILEPATHXL )
		os.system( "attrib +s +h /s \"{}/*.*\" && attrib +s +h \"{}\"".format( self.dest_dir, self.dest_dir ) )
		os.system( "attrib +s +h  \"{}\"".format( root_dir ) )
