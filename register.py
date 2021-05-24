import tkinter as tk
import tkinter.ttk as ttk
from openpyxl import *
import os, shutil
import string
import random
from tkinter import messagebox
import tempfile
import pyAesCrypt
from Gdrive import *


class Register:

	def __init__( self, master, service, home, homeId ):
		# build ui
		self.service = service
		self.home = home
		self.homeId = homeId
		frame_2 = ttk.Frame( master )
		user_L = ttk.Label( frame_2 )
		user_L.config( text='User Name' )
		user_L.grid( padx='5', pady='5' )
		self.user_E = ttk.Entry( frame_2 )
		self.user_E.config( font='{Arial} 12 {bold}', width='22' )
		self.user_E.grid( column='1', padx='5', pady='5', row='0' )
		password_L = ttk.Label( frame_2 )
		password_L.config( text='Password' )
		password_L.grid( column='0', padx='5', pady='5', row='1' )
		self.password_E = ttk.Entry( frame_2 )
		self.password_E.config( font='{Arial} 12 {bold}', width='22', show='•' )
		self.password_E.grid( column='1', padx='5', pady='5', row='1' )
		reenter_L = ttk.Label( frame_2 )
		reenter_L.config( text='Re Enter' )
		reenter_L.grid( column='0', padx='10', pady='5', row='2' )
		self.reenter_E = ttk.Entry( frame_2 )
		self.reenter_E.config( font='{Arial} 12 {bold}', width='22', show='•' )
		self.reenter_E.grid( column='1', padx='5', pady='5', row='2' )
		register_B = ttk.Button( frame_2 )
		register_B.config( default='normal', text='Register' )
		register_B.grid( column='0', padx='10', pady='20', row='3' )
		register_B.configure( command=self.register )
		cancel_B = ttk.Button( frame_2 )
		cancel_B.config( default='normal', text='Cancel' )
		cancel_B.grid( column='2', row='3' )
		cancel_B.configure( command=self.cancel )
		frame_2.config( height='200', width='200' )
		frame_2.place( anchor='nw', height='250', width='480', x='0', y='10' )

		# Main widget
		self.mainwindow = frame_2
		self.config()

	def config( self ):
		self.bufferSize = 64 * 1024
		# self.home = os.path.abspath( os.path.join( os.path.expanduser( '~' ), "AppData\\Local\\Programs\\Invisible" ) )
		# self.home = "Invisible"

	def random_str( self ):
		return ( ''.join( random.choices( string.ascii_uppercase + string.digits + string.ascii_lowercase, k=12 ) ) )

	def register( self ):
		self.user = self.user_E.get()
		self.password = self.password_E.get()
		self.reenter = self.reenter_E.get()
		self.userId = getFolderId( self.service, self.user, self.homeId )
		if self.userId:
			messagebox.showerror( "Exist", "User with this name already exists !" )
		elif ( self.user == "" or self.password == "" or self.reenter == "" ):
			messagebox.showerror( "Empty", "Empty Field !" )
		elif ( self.reenter_E.get() != self.password_E.get() ):
			messagebox.showerror( "Match", "Password Not Matched !" )
		else:
			# self.userdir = os.path.join( self.home, 'Users' )
			if not os.path.exists( os.path.join( self.home, self.user ) ):
				os.makedirs( os.path.join( self.home, self.user ) )
			self.userxl = os.path.join( self.home, self.user, self.user )
			# self.tempxl = tempfile.mkstemp( suffix='.xlsx' )[ 1 ]
			self.userId = createFolder( self.service, self.user, self.homeId )
			self.tempxl = os.path.join( tempfile.gettempdir(), self.random_str() + '.xlsx' )

			WB = Workbook()
			sheet = WB.active
			# dirname = self.random_str()

			dirlists = self.random_str()
			filelists = self.random_str()
			sheet.cell( row=1, column=1 ).value = self.user_E.get()
			sheet.cell( row=1, column=2 ).value = self.password_E.get()
			# sheet.cell( row=1, column=3 ).value = dirname
			sheet.cell( row=1, column=3 ).value = dirlists
			sheet.cell( row=1, column=4 ).value = filelists

			WB.save( self.tempxl )

			pyAesCrypt.encryptFile( self.tempxl, self.userxl, self.password, self.bufferSize )
			WB.close()
			fileId = uploadFile( self.service, self.user, self.userxl, self.userId )
			os.remove( self.tempxl )
			os.remove( self.userxl )
			# if not os.path.exists( os.path.join( self.home, dirname ) ):
			# 	os.makedirs( os.path.join( self.home, dirname ) )

			# if not os.path.exists( os.path.join( self.home, dirname, dirlists ) ):
			tempxl = os.path.join( tempfile.gettempdir(), self.random_str() + '.xlsx' )
			WB = Workbook()
			sheet = WB.active
			sheet.cell( row=1, column=1 ).value = "Folder Id"
			sheet.cell( row=1, column=2 ).value = "Folder Name"
			sheet.cell( row=1, column=3 ).value = "Temp Name"

			WB.save( tempxl )
			pyAesCrypt.encryptFile(
			    tempxl, os.path.join( self.home, self.user, dirlists ), self.password, self.bufferSize
			    )
			WB.close()
			os.remove( tempxl )
			fileId = uploadFile( self.service, dirlists, os.path.join( self.home, self.user, dirlists ), self.userId )
			os.remove( os.path.join( self.home, self.user, dirlists ) )

			messagebox.showinfo( "Done", "User {} Successfully Registerd !".format( self.user ) )
			self.mainwindow.destroy()

	def cancel( self ):
		self.mainwindow.destroy()

	def run( self ):
		self.mainwindow.mainloop()


if __name__ == '__main__':
	root = tk.Tk()
	app = Register( root )
	#print( app.random_str() )
	# app.run()
