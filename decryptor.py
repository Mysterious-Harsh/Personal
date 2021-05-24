import os
import string
import random
import shutil
import pyAesCrypt
import tempfile
from openpyxl import *


class Decryptor:

	def __init__( self, src_dir, dirlists, filelists, key ):
		self.key = key
		self.src_dir = src_dir
		# self.dest_dir = dest_dir
		self.bufferSize = 64 * 1024
		self.dirlists = dirlists
		self.filelists = filelists
		# password = "foopassword"

	def random_str( self, n ):
		return ( ''.join( random.choices( string.ascii_uppercase + string.digits + string.ascii_lowercase, k=n ) ) )

	def decrypt( self ):
		print( "yes" )

		FILEPATHXL = os.path.join( tempfile.gettempdir(), self.random_str( 12 ) + '.xlsx' )

		# self.Flag = False
		# print( os.path.join( self.src_dir, self.filelists ) )

		if os.path.exists( os.path.join( self.src_dir, self.filelists ) ):
			pyAesCrypt.decryptFile(
			    os.path.join( self.src_dir, self.filelists ), FILEPATHXL, self.key, self.bufferSize
			    )
			WB = load_workbook( FILEPATHXL )
			sheet = WB.active
			M_R = sheet.max_row

			for i in range( 1, M_R + 1 ):
				filepath = sheet.cell( row=i, column=1 ).value
				temp = sheet.cell( row=i, column=2 ).value
				if filepath != None and temp != None:
					DIR = os.path.dirname( filepath )

					if not os.path.exists( DIR ):
						os.makedirs( DIR )

					try:
						pyAesCrypt.decryptFile( temp, filepath, self.key, self.bufferSize )
					except:
						file, ext = os.path.splitext( filepath )
						file = file + "_" + self.random_str( 5 ) + ext
						pyAesCrypt.decryptFile( temp, file, self.key, self.bufferSize )

			for root, dirs, files in os.walk( self.src_dir ):
				for name in files:
					try:
						os.remove( os.path.join( root, name ) )
					except:
						pass
				for name in dirs:
					try:
						shutil.rmtree( os.path.join( root, name ) )
					except:
						pass
			try:
				shutil.rmtree( self.src_dir )

			except:
				pass
			os.remove( FILEPATHXL )
